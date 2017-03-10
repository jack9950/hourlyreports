import openpyxl
from openpyxl import Workbook
from get_pogo_sales_breakdown import get_pogo_sales_breakdown
from teams import agent_ids_to_names
import time
from datetime import datetime

homeFolder = 'C:\\Users\\Jackson.Ndiho\\Documents\\Sales\\'
pogoSalesReportLocation = homeFolder + 'bounce_energy_iqor_report_18.xls'

sales_numbers = get_pogo_sales_breakdown(pogoSalesReportLocation)
sales_numbers.sort()

# for item in sales_numbers:
#     print(item)

template = openpyxl.load_workbook(homeFolder + 'breakdown_test_template.xlsx')
template_sheets = template.get_sheet_names()
template_first_sheet = template.get_sheet_by_name(template_sheets[0])

#template_first_sheet["A5"].value = "testing"
currentDate = datetime.now().strftime("%A %m-%d-%Y")
template_first_sheet["A2"] = currentDate

row = 5
for sale in sales_numbers:
    template_first_sheet["A" + str(row)].value = sale[0]
    template_first_sheet["B" + str(row)].value = sale[1]
    template_first_sheet["C" + str(row)].value = sale[2]
    template_first_sheet["D" + str(row)].value = sale[3]
    row += 1


finalReportName = 'breakdown_test'
currentDate = datetime.now().strftime("%m%d%Y")
currentTime = time.strftime("%I%M%S%p")
#print(currentDate + "_" + currentTime)
template.save(homeFolder + finalReportName + "_" + currentDate + "_" + currentTime + ".xlsx")

# callsHandledReportLocation = homeFolder +'Bounce_Hourly_Sales_Report_03082017.xls'
# fcpReportLocation = homeFolder + 'HourlyProducts_Added.xls'
# DEPPreportLocation = homeFolder + 'products_sonar_03082017.xls'
# hiveNewServiceReportLocation = homeFolder + 'products_sonar_03082017.xls'
# hiveRenewalsReportLocation = homeFolder + 'hive_renewals_03082017.xls'

# agent_IDs_to_names = {2062026: "BECERRA, DOLORES", 2062062: "BROWN, ADRIANE"}
#
# ["BROWN, ADRIANNE", [2092985, 1443822, "Deposit due"], [2092021, 1444496, "Ercot/ISO Processing"] ]


# import time
# from datetime import datetime
# currentDate = datetime.now().strftime("%m%d%Y")
# currentTime = time.strftime("%I%M%S")
# print(currentDate + "_" + currentTime)

# # from get_HIVE_renewals import get_HIVE_renewals
# #
# # homeFolder = 'C:\\Users\\Jackson.Ndiho\\Documents\\Sales\\'
# #
# # callsHandledReportLocation = homeFolder +'Bounce_Hourly_Sales_Report_03082017.xls'
# # pogoSalesReportLocation = homeFolder + 'bounce_energy_iqor_report_18.xls'
# # fcpReportLocation = homeFolder + 'HourlyProducts_Added.xls'
# # #DEPPreportLocation = homeFolder + 'BounceEnergyProducts Added2017-03-08.xls'
# # DEPPreportLocation = homeFolder + 'products_sonar_03082017.xls'
# # hiveNewServiceReportLocation = homeFolder + 'products_sonar_03082017.xls'
# # hiveRenewalsReportLocation = homeFolder + 'hive_renewals_03082017.xls'
#
# # result = get_HIVE_renewals(hiveRenewalsReportLocation)
# #
# # for item in result:
# #     print(item)
#
# import openpyxl
# from openpyxl import Workbook
# from openpyxl.styles import Style, Font, Border, Side, Fill, PatternFill
# from get_agent_ids_and_calls import get_agent_ids_and_calls
# from get_pogo_sales import get_pogo_sales
# #from get_nest_sales import get_nest_sales
# from get_DEPP_sales import get_DEPP_sales
# from get_fcp_sales import get_fcp_sales
# from get_HIVE_new_service import get_HIVE_new_service
# from get_HIVE_renewals import get_HIVE_renewals
#
# wb = Workbook()
# ws = wb.active
#
# below_goal_text = "9C0006"
# below_goal_bg = "FFC7CE"
#
# close_to_goal_text = "9C6500"
# close_to_goal_bg = "FFEB9C"
#
# at_or_above_goal_text = "006100"
# at_or_above_goal_bg = "C6EFCE"
#
# myCell = ws['B5']
# myCell.value = "My Cell"
# myCell.font = Font(name='Calibri', size=11, bold=True)
# myCell.fill = PatternFill("solid", fgColor=agent_below_goal)
# # agent_below_goal = Style()
# # agent_below_goal.fill = PatternFill("solid", fgColor="ff6666")
#
# myCell.style = agent_below_goal
#print(agent_below_goal)

#wb.save("test.xlsx")
