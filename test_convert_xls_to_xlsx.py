import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

def open_xls_as_xlsx(filename):
# first open using xlrd    book = xlrd.open_workbook(filename)
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(1)
    nrows, ncols = sheet.nrows, sheet.ncols

    
    # print("Number of rows: ", sheet.nrows)
    # print("Number of cols: ", sheet.ncols)
    # print('inside converter: nrows: ', nrows, 'ncols: ', ncols)

    # prepare a xlsx sheet
    # book1 = Workbook()
    # sheet1 = book1.get_active_sheet()

    # for row in range(1, nrows+1):
    #     for col in range(1, ncols+1):
    #         sheet1.cell(row=row, column=col).value = sheet.cell_value(row-1, col-1)
    #         print(sheet1.cell(row=row, column=col).value)
    values = []
    # for row in range(0, nrows+1):
    #     for col in range(0, ncols+1):
    #         print(sheet.cell_value(row, col))
    for row in range(6, nrows):
        if sheet.cell_value(row,4) != '':
            values.append([sheet.cell_value(row,4), sheet.cell_value(row,11), sheet.cell_value(row,47)])
            #print(sheet.cell_value(row,4),": ", sheet.cell_value(row,47))

    # for item in values:
    #     print(item)

    return values